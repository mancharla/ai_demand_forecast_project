import io
import os
import json
import pandas as pd

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import PieChart, Reference

from app.database import get_db
from app.models import Forecast, Notification, Dataset
from app.utils.dependencies import get_current_user
from app.services.activity_logger import log_activity
from app.services.business_insights import generate_business_insights
from app.services.seasonal_trends import detect_seasonal_trends
from app.services.anomaly_detection import detect_sales_anomalies
from app.services.model_comparison import compare_forecasting_models
from app.services.email_service import send_email_notification

from app.services.professional_reports import (
    create_forecast_excel_report,
    create_analytics_excel_report,
    create_forecast_comparison_excel,
    create_dashboard_pdf_report,
)


router = APIRouter(
    prefix="/reports",
    tags=["Reports"],
)


def extract_forecast_items(forecast):
    try:
        value = json.loads(forecast.forecast_values or "{}")
    except Exception:
        value = {}

    if value.get("type") == "model_comparison_metric":
        return []

    product_forecasts = value.get("product_forecasts", [])

    if product_forecasts:
        items = []

        for item in product_forecasts:
            items.append(
                {
                    "product": item.get("product", "N/A"),
                    "predicted_sales": float(
                        item.get("predicted_sales", 0) or 0
                    ),
                    "model": item.get("model_used", forecast.model_name),
                    "dataset_id": forecast.dataset_id,
                    "created_at": forecast.created_at,
                }
            )

        return items

    product = (
        value.get("product")
        or value.get("product_name")
        or value.get("top_product")
        or value.get("top_demand_product")
        or "N/A"
    )

    predicted_sales = (
        value.get("predicted_sales")
        or value.get("sales")
        or value.get("forecast")
        or 0
    )

    return [
        {
            "product": product,
            "predicted_sales": float(predicted_sales or 0),
            "model": forecast.model_name,
            "dataset_id": forecast.dataset_id,
            "created_at": forecast.created_at,
        }
    ]


@router.get("/preview")
def report_preview(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    forecasts = (
        db.query(Forecast)
        .filter(
            Forecast.user_id == current_user.id,
            ~Forecast.forecast_values.like("%model_comparison_metric%"),
        )
        .order_by(Forecast.created_at.desc())
        .all()
    )

    forecast_data = []

    for forecast in forecasts:
        forecast_data.extend(extract_forecast_items(forecast))

    total_sales = sum(
        float(item["predicted_sales"] or 0)
        for item in forecast_data
    )

    unique_products = {
        item["product"]
        for item in forecast_data
        if item["product"] != "N/A"
    }

    return {
        "user_name": current_user.name,
        "total_sales": round(total_sales, 2),
        "total_products": len(unique_products),
        "forecast_count": len(forecasts),
        "forecast_data": forecast_data,
    }


@router.get("/forecast/excel")
def download_forecast_excel(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    forecasts = (
        db.query(Forecast)
        .filter(
            Forecast.user_id == current_user.id,
            ~Forecast.forecast_values.like("%model_comparison_metric%"),
        )
        .order_by(Forecast.created_at.desc())
        .all()
    )

    output = create_forecast_excel_report(forecasts)

    notification = Notification(
        user_id=current_user.id,
        message="Professional forecast Excel report generated successfully",
        type="success",
    )

    db.add(notification)

    send_email_notification(
        to_email=current_user.email,
        subject="Forecast Excel Report Ready",
        message=(
            f"Hello {current_user.name},\n\n"
            f"Your professional forecast Excel report has been generated successfully.\n\n"
            f"You can download it from the Reports section."
        ),
    )

    log_activity(
        db=db,
        user_id=current_user.id,
        action="FORECAST_EXCEL_DOWNLOAD",
        description="Downloaded professional forecast Excel report",
        module="Reports",
    )

    db.commit()

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                "attachment; filename=professional_forecast_report.xlsx"
            )
        },
    )


@router.get("/dashboard/pdf")
def download_dashboard_pdf(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    forecasts = (
        db.query(Forecast)
        .filter(
            Forecast.user_id == current_user.id,
            ~Forecast.forecast_values.like("%model_comparison_metric%"),
        )
        .order_by(Forecast.created_at.desc())
        .all()
    )

    output = create_dashboard_pdf_report(forecasts)

    notification = Notification(
        user_id=current_user.id,
        message="Professional dashboard PDF report generated successfully",
        type="success",
    )

    db.add(notification)

    send_email_notification(
        to_email=current_user.email,
        subject="Dashboard PDF Report Ready",
        message=(
            f"Hello {current_user.name},\n\n"
            f"Your dashboard PDF report has been generated successfully.\n\n"
            f"You can download it from the Reports section."
        ),
    )

    log_activity(
        db=db,
        user_id=current_user.id,
        action="DASHBOARD_PDF_DOWNLOAD",
        description="Downloaded professional dashboard PDF report",
        module="Reports",
    )

    db.commit()

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={
            "Content-Disposition": (
                "attachment; filename=professional_dashboard_report.pdf"
            )
        },
    )


@router.get("/analytics/summary/excel")
def download_analytics_summary_excel(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    datasets = (
        db.query(Dataset)
        .filter(Dataset.user_id == current_user.id)
        .all()
    )

    all_frames = []

    for dataset in datasets:
        file_path = os.path.join("uploads", dataset.filename)

        if not os.path.exists(file_path):
            continue

        if dataset.filename.endswith(".csv"):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)

        all_frames.append(df)

    if not all_frames:
        raise HTTPException(
            status_code=404,
            detail="No dataset available for analytics report",
        )

    df = pd.concat(all_frames, ignore_index=True)

    if "Sales" in df.columns:
        df["Sales"] = pd.to_numeric(
            df["Sales"],
            errors="coerce",
        ).fillna(0)

    insights = generate_business_insights(df)
    seasonal_data = detect_seasonal_trends(df)
    anomalies = detect_sales_anomalies(df)

    output = create_analytics_excel_report(
        df=df,
        insights=insights,
        seasonal_data=seasonal_data,
        anomalies=anomalies,
    )

    notification = Notification(
        user_id=current_user.id,
        message="Professional analytics summary Excel generated successfully",
        type="success",
    )

    db.add(notification)

    send_email_notification(
        to_email=current_user.email,
        subject="Analytics Summary Report Ready",
        message=(
            f"Hello {current_user.name},\n\n"
            f"Your analytics summary Excel report has been generated successfully.\n\n"
            f"You can download it from the Reports section."
        ),
    )

    log_activity(
        db=db,
        user_id=current_user.id,
        action="ANALYTICS_SUMMARY_DOWNLOAD",
        description="Downloaded professional analytics summary Excel report",
        module="Reports",
    )

    db.commit()

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                "attachment; filename=professional_analytics_summary.xlsx"
            )
        },
    )


@router.get("/forecast-comparison/{dataset_id}/excel")
def download_forecast_comparison_excel(
    dataset_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    dataset = (
        db.query(Dataset)
        .filter(
            Dataset.id == dataset_id,
            Dataset.user_id == current_user.id,
        )
        .first()
    )

    if not dataset:
        raise HTTPException(
            status_code=404,
            detail="Dataset not found",
        )

    file_path = os.path.join("uploads", dataset.filename)

    if not os.path.exists(file_path):
        raise HTTPException(
            status_code=404,
            detail="Dataset file not found",
        )

    if dataset.filename.endswith(".csv"):
        df = pd.read_csv(file_path)
    else:
        df = pd.read_excel(file_path)

    result = compare_forecasting_models(df)

    output = create_forecast_comparison_excel(
        result=result,
        dataset_id=dataset_id,
    )

    notification = Notification(
        user_id=current_user.id,
        message="Forecast comparison Excel generated successfully",
        type="success",
    )

    db.add(notification)

    send_email_notification(
        to_email=current_user.email,
        subject="Forecast Comparison Report Ready",
        message=(
            f"Hello {current_user.name},\n\n"
            f"Your forecast comparison Excel report has been generated successfully.\n\n"
            f"You can download it from the Reports section."
        ),
    )

    log_activity(
        db=db,
        user_id=current_user.id,
        action="FORECAST_COMPARISON_EXCEL_DOWNLOAD",
        description=f"Downloaded forecast comparison report for dataset {dataset_id}",
        module="Reports",
    )

    db.commit()

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                f"attachment; filename=forecast_comparison_dataset_{dataset_id}.xlsx"
            )
        },
    )


@router.get("/dashboard/summary/excel")
def download_dashboard_summary_excel(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    forecasts = (
        db.query(Forecast)
        .filter(
            Forecast.user_id == current_user.id,
            ~Forecast.forecast_values.like("%model_comparison_metric%"),
        )
        .order_by(Forecast.id.desc())
        .all()
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dashboard Summary"

    sheet.append(["Metric", "Value"])

    total_forecasts = len(forecasts)
    total_sales = 0
    models = {}

    for forecast in forecasts:
        items = extract_forecast_items(forecast)

        for item in items:
            total_sales += float(item["predicted_sales"] or 0)

        models[forecast.model_name] = models.get(forecast.model_name, 0) + 1

    sheet.append(["Total Forecasts", total_forecasts])
    sheet.append(["Total Predicted Sales", round(total_sales, 2)])
    sheet.append(["Models Used", ", ".join(models.keys())])
    sheet.append(["Generated By", current_user.name])

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    chart_sheet = workbook.create_sheet("Model Usage")
    chart_sheet.append(["Model", "Count"])

    for model, count in models.items():
        chart_sheet.append([model, count])

    if models:
        pie = PieChart()
        pie.title = "Model Usage Distribution"

        labels = Reference(
            chart_sheet,
            min_col=1,
            min_row=2,
            max_row=len(models) + 1,
        )

        data = Reference(
            chart_sheet,
            min_col=2,
            min_row=1,
            max_row=len(models) + 1,
        )

        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 8
        pie.width = 12

        chart_sheet.add_chart(pie, "D2")

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                "attachment; filename=dashboard_summary.xlsx"
            )
        },
    )