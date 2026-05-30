import json
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics import renderPDF
from openpyxl.cell.cell import MergedCell


def style_header(sheet):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

def auto_width(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = None

        for cell in column:
            if isinstance(cell, MergedCell):
                continue

            if column_letter is None:
                column_letter = cell.column_letter

            try:
                if cell.value:
                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )
            except Exception:
                pass

        if column_letter:
            sheet.column_dimensions[column_letter].width = (
                max_length + 4
            )
def create_kpi_card(sheet, cell, title, value, color):
    sheet[cell] = title
    sheet[cell].font = Font(bold=True, color="FFFFFF")
    sheet[cell].fill = PatternFill("solid", fgColor=color)
    sheet[cell].alignment = Alignment(horizontal="center")

    value_cell = cell[0] + str(int(cell[1:]) + 1)
    sheet[value_cell] = value
    sheet[value_cell].font = Font(bold=True, size=16, color="000000")
    sheet[value_cell].alignment = Alignment(horizontal="center")
    sheet[value_cell].fill = PatternFill("solid", fgColor="EAF2F8")


def create_forecast_excel_report(forecasts):
    total_predicted_sales = 0
    model_counts = {}
    forecast_rows = []
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    summary_sheet.merge_cells("A1:H1")

    summary_sheet["A1"] = "AI Demand Forecasting - Forecast Report Summary"
    summary_sheet["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    summary_sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary_sheet["A1"].alignment = Alignment(horizontal="center")

    # Process forecasts to populate forecast_rows first
    for forecast in forecasts:
        try:
            value = json.loads(forecast.forecast_values or "{}")
        except Exception:
            value = {}

        product = value.get("product", "N/A")
        predicted_sales = float(value.get("predicted_sales", 0) or 0)
        model_name = forecast.model_name or "N/A"

        forecast_rows.append([
            forecast.id,
            forecast.dataset_id,
            product,
            predicted_sales,
            model_name,
            str(forecast.created_at),
        ])

        total_predicted_sales += predicted_sales

        if model_name not in model_counts:
            model_counts[model_name] = 0
        model_counts[model_name] += 1

    # Now add KPI cards with calculated values
    avg_sales = (
        round(total_predicted_sales / len(forecast_rows), 2)
        if forecast_rows else 0
    )
    max_sales = (
        max(row[3] for row in forecast_rows)
        if forecast_rows else 0
    )

    create_kpi_card(summary_sheet, "A3", "Total Forecasts", len(forecast_rows), "2563EB")
    create_kpi_card(summary_sheet, "C3", "Total Predicted Sales", round(total_predicted_sales, 2), "059669")
    create_kpi_card(summary_sheet, "E3", "Average Sales", avg_sales, "7C3AED")
    create_kpi_card(summary_sheet, "G3", "Max Sales", round(max_sales, 2), "DC2626")

    style_header(summary_sheet)
    auto_width(summary_sheet)

    forecast_sheet = workbook.create_sheet("Forecast Details")
    forecast_sheet.append([
        "Forecast ID",
        "Dataset ID",
        "Product",
        "Predicted Sales",
        "Model",
        "Created At",
    ])

    for row in forecast_rows:
        forecast_sheet.append(row)

    style_header(forecast_sheet)
    auto_width(forecast_sheet)

    if len(forecast_rows) > 0:
        chart = BarChart()
        chart.title = "Predicted Sales by Product"
        chart.y_axis.title = "Predicted Sales"
        chart.x_axis.title = "Product"

        data = Reference(
            forecast_sheet,
            min_col=4,
            min_row=1,
            max_row=min(len(forecast_rows) + 1, 15),
        )

        categories = Reference(
            forecast_sheet,
            min_col=3,
            min_row=2,
            max_row=min(len(forecast_rows) + 1, 15),
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 10
        chart.width = 20

        forecast_sheet.add_chart(chart, "H2")

    model_sheet = workbook.create_sheet("Model Usage")
    model_sheet.append(["Model", "Count"])

    for model, count in model_counts.items():
        model_sheet.append([model, count])

    style_header(model_sheet)
    auto_width(model_sheet)

    if model_counts:
        pie = PieChart()
        pie.title = "Model Usage Distribution"

        labels = Reference(
            model_sheet,
            min_col=1,
            min_row=2,
            max_row=len(model_counts) + 1,
        )

        data = Reference(
            model_sheet,
            min_col=2,
            min_row=1,
            max_row=len(model_counts) + 1,
        )

        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 8
        pie.width = 12

        model_sheet.add_chart(pie, "D2")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


def create_analytics_excel_report(df, insights, seasonal_data, anomalies):
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Executive Summary"

    total_sales = float(df["Sales"].sum()) if "Sales" in df.columns else 0
    total_products = df["Product"].nunique() if "Product" in df.columns else 0
    total_regions = df["Region"].nunique() if "Region" in df.columns else 0
    total_categories = df["Category"].nunique() if "Category" in df.columns else 0

    summary_sheet.merge_cells("A1:H1")
    summary_sheet["A1"] = "AI Demand Forecasting - Analytics Executive Summary"
    summary_sheet["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    summary_sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary_sheet["A1"].alignment = Alignment(horizontal="center")

    summary_sheet.merge_cells("A2:H2")
    summary_sheet["A2"] = "Region, Category, Seasonal Trend, Revenue, AI Insights and Risk Analysis"
    summary_sheet["A2"].font = Font(size=11, italic=True, color="666666")
    summary_sheet["A2"].alignment = Alignment(horizontal="center")

    create_kpi_card(summary_sheet, "A4", "Total Records", len(df), "2563EB")
    create_kpi_card(summary_sheet, "C4", "Total Sales", round(total_sales, 2), "059669")
    create_kpi_card(summary_sheet, "E4", "Products", total_products, "7C3AED")
    create_kpi_card(summary_sheet, "G4", "Anomalies", len(anomalies), "DC2626")

    summary_sheet["A8"] = "Executive Summary"
    summary_sheet["A8"].font = Font(bold=True, size=14)

    summary_sheet["A9"] = (
        "This analytics summary provides a business-level view of demand performance, "
        "regional contribution, category sales, seasonal trend behavior, anomaly risk, "
        "and AI-generated business insights."
    )

    summary_sheet["A11"] = "Business Recommendations"
    summary_sheet["A11"].font = Font(bold=True, size=14)

    recommendations = [
        "Increase monitoring for products and regions showing unusual sales patterns.",
        "Use region-wise demand data to plan inventory allocation.",
        "Review category-level insights before pricing and procurement decisions.",
        "Use seasonal trends to prepare demand planning before peak months.",
    ]

    row = 12
    for rec in recommendations:
        summary_sheet[f"A{row}"] = f"• {rec}"
        row += 1

    auto_width(summary_sheet)

    if "Region" in df.columns and "Sales" in df.columns:
        region_sheet = workbook.create_sheet("Region Analytics")
        region_sheet.append(["Region", "Total Sales"])

        region_data = df.groupby("Region")["Sales"].sum().reset_index()

        for _, row in region_data.iterrows():
            region_sheet.append([row["Region"], float(row["Sales"])])

        style_header(region_sheet)
        auto_width(region_sheet)

        chart = BarChart()
        chart.title = "Region Wise Sales"
        chart.y_axis.title = "Sales"
        chart.x_axis.title = "Region"

        data = Reference(region_sheet, min_col=2, min_row=1, max_row=len(region_data) + 1)
        cats = Reference(region_sheet, min_col=1, min_row=2, max_row=len(region_data) + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 18

        region_sheet.add_chart(chart, "D2")

    if "Category" in df.columns and "Sales" in df.columns:
        category_sheet = workbook.create_sheet("Category Insights")
        category_sheet.append(["Category", "Total Sales", "Average Sales", "Records"])

        category_data = df.groupby("Category")["Sales"].agg(["sum", "mean", "count"]).reset_index()

        for _, row in category_data.iterrows():
            category_sheet.append([
                row["Category"],
                float(row["sum"]),
                round(float(row["mean"]), 2),
                int(row["count"]),
            ])

        style_header(category_sheet)
        auto_width(category_sheet)

        pie = PieChart()
        pie.title = "Category Sales Distribution"

        labels = Reference(category_sheet, min_col=1, min_row=2, max_row=len(category_data) + 1)
        data = Reference(category_sheet, min_col=2, min_row=1, max_row=len(category_data) + 1)

        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 8
        pie.width = 12

        category_sheet.add_chart(pie, "F2")

    trend_sheet = workbook.create_sheet("Seasonal Trends")
    trend_sheet.append(["Type", "Name", "Sales"])

    for item in seasonal_data.get("monthly_trends", []):
        trend_sheet.append(["Month", item["month"], item["sales"]])

    for item in seasonal_data.get("quarterly_trends", []):
        trend_sheet.append(["Quarter", item["quarter"], item["sales"]])

    style_header(trend_sheet)
    auto_width(trend_sheet)

    if len(seasonal_data.get("monthly_trends", [])) > 0:
        line = LineChart()
        line.title = "Monthly Sales Trend"
        line.y_axis.title = "Sales"
        line.x_axis.title = "Month"

        month_count = len(seasonal_data.get("monthly_trends", []))

        data = Reference(trend_sheet, min_col=3, min_row=1, max_row=month_count + 1)
        cats = Reference(trend_sheet, min_col=2, min_row=2, max_row=month_count + 1)

        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.height = 10
        line.width = 18

        trend_sheet.add_chart(line, "E2")

    insight_sheet = workbook.create_sheet("AI Insights")
    insight_sheet.append(["Business Insight"])

    for insight in insights:
        insight_sheet.append([insight])

    style_header(insight_sheet)
    auto_width(insight_sheet)

    anomaly_sheet = workbook.create_sheet("Anomalies")
    anomaly_sheet.append(["Date", "Product", "Region", "Sales", "Reason"])

    for item in anomalies:
        anomaly_sheet.append([
            item.get("date"),
            item.get("product"),
            item.get("region"),
            item.get("sales"),
            item.get("reason"),
        ])

    style_header(anomaly_sheet)
    auto_width(anomaly_sheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


def create_forecast_comparison_excel(result, dataset_id):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Model Comparison"

    sheet.append(["Dataset ID", dataset_id])
    sheet.append([])
    sheet.append(["Model", "MAE", "RMSE"])

    for model in result.get("all_models", []):
        sheet.append([
            model.get("model"),
            model.get("mae"),
            model.get("rmse"),
        ])

    style_header(sheet)
    auto_width(sheet)

    if result.get("all_models"):
        chart = BarChart()
        chart.title = "Model RMSE Comparison"
        chart.y_axis.title = "RMSE"
        chart.x_axis.title = "Model"

        start_row = 4
        end_row = start_row + len(result.get("all_models", [])) - 1

        data = Reference(sheet, min_col=3, min_row=3, max_row=end_row)
        cats = Reference(sheet, min_col=1, min_row=4, max_row=end_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 16

        sheet.add_chart(chart, "E4")

    summary = workbook.create_sheet("Best Model")
    summary.append(["Metric", "Value"])
    summary.append(["Best Model", result.get("best_model", {}).get("model")])
    summary.append(["Best MAE", result.get("best_model", {}).get("mae")])
    summary.append(["Best RMSE", result.get("best_model", {}).get("rmse")])

    style_header(summary)
    auto_width(summary)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


def create_dashboard_pdf_report(forecasts):
    buffer = BytesIO()

    pdf = canvas.Canvas(buffer, pagesize=letter)

    width, height = letter

    # ===== HEADER =====
    pdf.setFillColor(colors.HexColor("#0F172A"))
    pdf.rect(0, height - 90, width, 90, fill=1)

    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 24)
    pdf.drawString(40, height - 45, "AI Demand Forecasting Report")

    pdf.setFont("Helvetica", 12)
    pdf.drawString(
        40,
        height - 68,
        "Enterprise Forecast Analytics Dashboard",
    )

    # ===== FORECAST DATA =====
    rows = []

    for forecast in forecasts[:15]:
        try:
            value = json.loads(forecast.forecast_values or "{}")
        except Exception:
            value = {}

        rows.append([
            value.get("product", "N/A"),
            float(value.get("predicted_sales", 0) or 0),
            forecast.model_name,
        ])

    total_sales = sum(row[1] for row in rows)

    avg_sales = (
        round(total_sales / len(rows), 2)
        if rows else 0
    )

    max_sales = (
        max(row[1] for row in rows)
        if rows else 0
    )

    # ===== KPI CARDS =====
    card_y = height - 170

    cards = [
        ("Forecast Records", len(rows), "#2563EB"),
        ("Total Predicted Sales", round(total_sales, 2), "#059669"),
        ("Average Sales", avg_sales, "#7C3AED"),
        ("Max Sales", round(max_sales, 2), "#EA580C"),
    ]

    x = 40

    for title, value, color in cards:
        pdf.setFillColor(colors.HexColor(color))
        pdf.roundRect(x, card_y, 120, 70, 10, fill=1)

        pdf.setFillColor(colors.white)

        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(x + 10, card_y + 50, str(title))

        pdf.setFont("Helvetica-Bold", 18)
        pdf.drawString(x + 10, card_y + 25, str(value))

        x += 135

    # ===== TABLE =====
    table_y = height - 290

    pdf.setFillColor(colors.HexColor("#1E293B"))
    pdf.roundRect(40, table_y + 15, 520, 28, 6, fill=1)

    pdf.setFillColor(colors.white)

    pdf.setFont("Helvetica-Bold", 11)

    pdf.drawString(55, table_y + 25, "Product")
    pdf.drawString(240, table_y + 25, "Predicted Sales")
    pdf.drawString(420, table_y + 25, "Model")

    y = table_y - 5

    for row in rows[:10]:
        pdf.setFillColor(colors.black)

        pdf.setFont("Helvetica", 10)

        pdf.drawString(55, y, str(row[0])[:25])
        pdf.drawString(260, y, str(round(row[1], 2)))
        pdf.drawString(420, y, str(row[2])[:20])

        pdf.line(45, y - 8, 550, y - 8)

        y -= 22

    # ===== BAR CHART =====
    if rows:
        drawing = Drawing(500, 260)

        chart = VerticalBarChart()

        chart.x = 50
        chart.y = 40
        chart.height = 180
        chart.width = 380

        chart.data = [[row[1] for row in rows[:8]]]

        chart.categoryAxis.categoryNames = [
            str(row[0])[:8]
            for row in rows[:8]
        ]

        chart.valueAxis.valueMin = 0

        chart.bars[0].fillColor = colors.HexColor("#2563EB")

        chart.categoryAxis.labels.angle = 20

        drawing.add(chart)

        renderPDF.draw(
            drawing,
            pdf,
            40,
            120,
        )

    # ===== INSIGHTS =====
    pdf.setFont("Helvetica-Bold", 15)

    pdf.drawString(
        40,
        95,
        "AI Business Insights",
    )

    insights = [
        "• High demand products show strong seasonal patterns.",
        "• Random Forest model generated highest forecasting accuracy.",
        "• Electronics category dominates projected sales trends.",
        "• Inventory optimization recommended for top-selling products.",
    ]

    pdf.setFont("Helvetica", 11)

    insight_y = 75

    for insight in insights:
        pdf.drawString(50, insight_y, insight)
        insight_y -= 18

    # ===== FOOTER =====
    pdf.setFillColor(colors.HexColor("#64748B"))

    pdf.setFont("Helvetica", 9)

    pdf.drawString(
        40,
        25,
        "Generated by AI Demand Forecasting Platform",
    )

    pdf.drawRightString(
        width - 40,
        25,
        "Confidential Analytics Report",
    )

    pdf.showPage()

    pdf.save()

    buffer.seek(0)

    return buffer